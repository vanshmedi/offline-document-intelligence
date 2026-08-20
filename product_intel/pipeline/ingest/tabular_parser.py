"""
CSV / XLSX parsing.

Price files and ERP extracts are row-per-SKU, so each row becomes its own
keyvalue fragment carrying a row locator. These are machine-authored feeds and
are marked STRUCTURED_FEED, which gives them a high reliability weight in
confidence scoring and high precedence in golden-record arbitration.
"""

from __future__ import annotations

import csv
import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Tuple

from product_intel.models import ExtractionMethod, Fragment

log = logging.getLogger(__name__)

MAX_ROWS = 100_000


def _clean(v: Any) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _read_csv(path: Path) -> Tuple[List[str], List[List[str]]]:
    with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
        sample = f.read(8192)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(f, dialect)
        rows = [[_clean(c) for c in row] for row in reader]
    if not rows:
        return [], []
    return rows[0], rows[1:]


def _read_xlsx(path: Path) -> Tuple[List[str], List[List[str]]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [[_clean(c) for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()
    rows = [r for r in rows if any(r)]
    if not rows:
        return [], []
    return rows[0], rows[1:]


def parse_tabular(path: Path, source_id: str) -> Tuple[List[Fragment], str, Dict[str, Any]]:
    """Parse a CSV/TSV/XLSX feed into one fragment per data row."""
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        header, data_rows = _read_xlsx(path)
    else:
        header, data_rows = _read_csv(path)

    stats: Dict[str, Any] = {"rows": len(data_rows), "columns": len(header), "warnings": []}
    if not header:
        raise ValueError(f"{path.name}: no header row found")
    if len(data_rows) > MAX_ROWS:
        stats["warnings"].append(f"truncated to first {MAX_ROWS} rows of {len(data_rows)}")
        data_rows = data_rows[:MAX_ROWS]

    fragments: List[Fragment] = []
    mirror: List[str] = [f"# {path.name}\n\nColumns: {', '.join(header)}\n\n"]

    for r_idx, row in enumerate(data_rows, start=2):  # row 1 is the header
        pairs = [[header[i], row[i]] for i in range(min(len(header), len(row))) if row[i]]
        if not pairs:
            continue
        text = "\n".join(f"{k}: {v}" for k, v in pairs)
        fragments.append(
            Fragment(
                fragment_id=f"{source_id}_r{r_idx}",
                source_id=source_id,
                kind="keyvalue",
                locator=f"row {r_idx}",
                text=text,
                table=pairs,
                method=ExtractionMethod.STRUCTURED_FEED,
                metadata={"row_number": r_idx, "row_scoped": True},
            )
        )
        mirror.append(f"\n### Row {r_idx}\n\n")
        mirror.extend(f"- **{k}**: {v}\n" for k, v in pairs)

    stats["fragments"] = len(fragments)
    return fragments, "".join(mirror), stats
